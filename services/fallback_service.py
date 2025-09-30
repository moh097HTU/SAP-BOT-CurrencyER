# services/fallback_service.py
from __future__ import annotations

import json
import logging
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Tuple

from services.config import config
from services.driver import get_driver
from services.auth import login
from services.ui import wait_for_shell_home, wait_ui5_idle
from pages.Shell.Search.element import ShellSearch
from pages.CurrencyExchangeRates.page import CurrencyExchangeRatesPage
from pages.CurrencyExchangeRates.elements.DraftFinder import DraftFinder
from pages.CurrencyExchangeRates.elements.ExcelExport.element import ExcelExporter

log = logging.getLogger("sapbot")

# ---------- formatting helpers ----------
def _ddmmyyyy(d: date) -> str:
    return f"{d.day:02d}.{d.month:02d}.{d.year:04d}"

def _iso(d: date) -> str:
    return d.strftime("%Y-%m-%d")

def _daterange_inclusive(d0: date, d1: date):
    cur = min(d0, d1)
    end = max(d0, d1)
    while cur <= end:
        yield cur
        cur = cur + timedelta(days=1)

# ---------- IO helpers ----------
BASE_DATA_DIR = Path("WebService") / "data"
FALLBACK_TRACK_DIR = Path("WebService") / "TrackDrivers" / "Fallback"
TEMP_DL_DIR = Path(config().get("REPORTS_DIR") or "reports") / "tmp_downloads"
TEMP_DL_DIR.mkdir(parents=True, exist_ok=True)
FALLBACK_TRACK_DIR.mkdir(parents=True, exist_ok=True)

# ---------- comparison helpers ----------
def _q_norm(q: str | None) -> str:
    s = (q or "").strip().lower()
    return "Indirect" if s.startswith("ind") else "Direct"

def _key_tuple(r: Dict[str, Any]) -> tuple:
    return (
        (r.get("ExchangeRateType") or "").strip().upper(),
        (r.get("ValidFrom") or "").strip(),
        (r.get("FromCurrency") or "").strip().upper(),
        (r.get("ToCurrency") or "").strip().upper(),
        _q_norm(r.get("Quotation")),
    )

def _rev_flip_tuple(r: Dict[str, Any]) -> tuple:
    # reverse currencies, flip quotation
    q = _q_norm(r.get("Quotation"))
    flipped = "Direct" if q == "Indirect" else "Indirect"
    return (
        (r.get("ExchangeRateType") or "").strip().upper(),
        (r.get("ValidFrom") or "").strip(),
        (r.get("ToCurrency") or "").strip().upper(),
        (r.get("FromCurrency") or "").strip().upper(),
        flipped,
    )

def _rev_same_tuple(r: Dict[str, Any]) -> tuple:
    # reverse currencies, keep quotation (covers SAP exports that already present reverse-side)
    return (
        (r.get("ExchangeRateType") or "").strip().upper(),
        (r.get("ValidFrom") or "").strip(),
        (r.get("ToCurrency") or "").strip().upper(),
        (r.get("FromCurrency") or "").strip().upper(),
        _q_norm(r.get("Quotation")),
    )

def _json_missing_vs_excel(excel_rows: list[dict], json_rows: list[dict]) -> list[dict]:
    def _q_norm(q: str) -> str:
        q = (q or "").strip().lower()
        return "Indirect" if q.startswith("ind") else "Direct"

    def _key_tuple(r: dict) -> tuple:
        # exact-direction key
        return (
            (r.get("ExchangeRateType") or "").strip().upper(),
            (r.get("ValidFrom") or "").strip(),
            (r.get("FromCurrency") or "").strip().upper(),
            (r.get("ToCurrency") or "").strip().upper(),
            _q_norm(r.get("Quotation")),
        )

    def _rev_flip_tuple(r: dict) -> tuple:
        # opposite direction + flipped quotation (the only valid equivalence)
        q = _q_norm(r.get("Quotation"))
        flipped_q = "Direct" if q == "Indirect" else "Indirect"
        return (
            (r.get("ExchangeRateType") or "").strip().upper(),
            (r.get("ValidFrom") or "").strip(),
            (r.get("ToCurrency") or "").strip().upper(),
            (r.get("FromCurrency") or "").strip().upper(),
            flipped_q,
        )

    # Build presence set from Excel (exact + reverse+flipped only)
    excel_keys: set[tuple] = set()
    for er in excel_rows or []:
        try:
            excel_keys.add(_key_tuple(er))
            excel_keys.add(_rev_flip_tuple(er))
        except Exception:
            continue

    # Anything in JSON not covered by those keys is missing
    missing: list[dict] = []
    for r in json_rows or []:
        k_exact = _key_tuple(r)
        k_rev   = _rev_flip_tuple(r)
        if k_exact in excel_keys or k_rev in excel_keys:
            continue
        missing.append({
            "ExchangeRateType": (r.get("ExchangeRateType") or "").strip().upper(),
            "FromCurrency": (r.get("FromCurrency") or "").strip().upper(),
            "ToCurrency": (r.get("ToCurrency") or "").strip().upper(),
            "ValidFrom": (r.get("ValidFrom") or "").strip(),
            "Quotation": _q_norm(r.get("Quotation")),
            "ExchangeRate": r.get("ExchangeRate"),
        })
    return missing

def _read_json_payload(day_iso: str) -> Tuple[List[Dict[str, Any]], bool]:
    p = BASE_DATA_DIR / day_iso / "exchange_rates_payload.json"
    if not p.exists():
        return [], False
    try:
        rows = json.loads(p.read_text(encoding="utf-8"))
        out: List[Dict[str, Any]] = [r for r in (rows or []) if isinstance(r, dict)]
        return out, True
    except Exception:
        return [], True

def _write_missing_tracker(day_iso: str, rows: List[Dict[str, Any]]) -> Path:
    out_path = FALLBACK_TRACK_DIR / f"{day_iso}.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")
    return out_path

# ---------- Excel parsing ----------
def _parse_rate(val: Any) -> float | None:
    """
    Accept numbers or numeric strings like '1,234.56789'. Return float or None.
    """
    if val is None:
        return None
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace(",", "")
        return float(s) if s else None
    except Exception:
        return None

def _read_excel_rows(xlsx_path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Parse SAP ListReport export and normalize rows.
    IMPORTANT CHANGE: a row counts if it has type/date/from/to; quotation defaults to Direct;
    rate is optional (may be None). This fixes undercounting (e.g., 378 vs 380).
    """
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required for Excel fallback parsing; install it via 'pip install openpyxl'.") from exc

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_cells = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = {(str(v).strip() if v is not None else ""): i for i, v in enumerate(header_cells)}

    def get(row_vals, name, *alts):
        for key in (name, *alts):
            if key in headers:
                return row_vals[headers[key]]
        return None

    rows: List[Dict[str, Any]] = []
    skipped_reasons: Dict[str, int] = {}

    for r in ws.iter_rows(min_row=2, values_only=True):
        exch_lbl = str(get(r, "Exchange Rate Type") or "").strip()
        exch_code = exch_lbl[:1].upper() if exch_lbl else ""

        vf_raw = get(r, "Valid From")
        if hasattr(vf_raw, "strftime"):
            valid_from = vf_raw.strftime("%d.%m.%Y")
        else:
            valid_from = str(vf_raw or "").strip()

        # Prefer Currency Pair, else From/To
        frm, to = "", ""
        cp = get(r, "Currency Pair")
        if cp:
            cp = str(cp).strip()
            frm, to = (cp.split("/", 1) + [""])[:2] if "/" in cp else (cp, "")
            frm, to = (frm or "").strip(), (to or "").strip()
        else:
            frm = str(get(r, "From Currency") or "").strip()
            to  = str(get(r, "To Currency") or "").strip()

        # Normalize currency cells like "AED (United ...)" -> "AED"
        if frm:
            frm = frm.split(" ", 1)[0].split("(")[0].strip().upper()
        if to:
            to = to.split(" ", 1)[0].split("(")[0].strip().upper()

        qtn_raw = str(get(r, "Quotation") or "").strip()
        quotation = "Indirect" if qtn_raw.lower().startswith("ind") else "Direct"  # default to Direct

        rate_val = get(r, "Rate 1:1", "Rate (1:1)")
        rate = _parse_rate(rate_val)  # optional

        # NEW relaxed keep-condition: only key fields are required
        if not exch_code:
            skipped_reasons["no_type"] = skipped_reasons.get("no_type", 0) + 1
            continue
        if not valid_from:
            skipped_reasons["no_date"] = skipped_reasons.get("no_date", 0) + 1
            continue
        if not frm or not to:
            skipped_reasons["no_pair"] = skipped_reasons.get("no_pair", 0) + 1
            continue

        rows.append({
            "ExchangeRateType": exch_code,
            "ValidFrom": valid_from,
            "FromCurrency": frm,
            "ToCurrency": to,
            "Quotation": quotation,
            "ExchangeRate": rate,  # may be None
        })

    diag = {
        "sheet_title": ws.title,
        "n_cols": len(header_cells),
        "headers": list(headers.keys()),
        "n_rows_parsed": len(rows),
        "file_bytes": xlsx_path.stat().st_size if xlsx_path.exists() else 0,
        "skipped_reasons": skipped_reasons,
    }
    return rows, diag

# ---------- main runner ----------
def run_collect_missing_range(day_from: date, day_to: date) -> Dict[str, Any]:
    """
    For each day in [day_from, day_to]:
      - open 'Currency Exchange Rates' app
      - set date filter, export to Excel
      - parse xlsx + read JSON WebService/data/<YYYY-MM-DD>/exchange_rates_payload.json
      - find JSON items missing from Excel (considering reverse pairs/quotation flip)
      - write ALWAYS a tracker at WebService/TrackDrivers/Fallback/<YYYY-MM-DD>.json
    """
    cfg = config()
    drv = None

    out: Dict[str, Any] = {
        "ok": True,
        "total_days": 0,
        "processed": 0,
        "total_missing": 0,
        "per_day": [],
    }

    try:
        TEMP_DL_DIR.mkdir(parents=True, exist_ok=True)

        drv = get_driver(headless=cfg["HEADLESS"], download_dir=str(TEMP_DL_DIR))
        login(drv)
        wait_for_shell_home(drv, timeout=cfg["EXPLICIT_WAIT_SECONDS"])
        wait_ui5_idle(drv, timeout=cfg["EXPLICIT_WAIT_SECONDS"])

        ShellSearch(drv).open_search().type_and_choose_app("Currency Exchange Rates")
        wait_ui5_idle(drv, timeout=30)
        CurrencyExchangeRatesPage(drv).ensure_in_app(max_attempts=3, settle_each=8)

        finder = DraftFinder(drv)
        exporter = ExcelExporter(drv)

        for d in _daterange_inclusive(day_from, day_to):
            out["total_days"] += 1
            day_iso = _iso(d)
            day_ddmmyyyy = _ddmmyyyy(d)

            ok = finder.set_effective_date_and_apply(day_ddmmyyyy, timeout=20)
            if not ok:
                # Write empty tracker anyway so you have a breadcrumb
                tracker_path = str(_write_missing_tracker(day_iso, []))
                out["per_day"].append({
                    "date": day_iso, "ok": False, "why": "date_set_failed",
                    "excel_rows": 0, "json_rows": 0, "missing": 0,
                    "tracker_path": tracker_path, "export_clicked": False,
                    "xlsx_path": "", "xlsx_size": 0, "headers_seen": [],
                    "json_file_exists": False,
                })
                continue

            xlsx_path, xlsx_size = exporter.export_now(download_dir=TEMP_DL_DIR, timeout=90)
            if not xlsx_path or not xlsx_path.exists() or xlsx_size == 0:
                tracker_path = str(_write_missing_tracker(day_iso, []))
                out["per_day"].append({
                    "date": day_iso, "ok": False, "why": "excel_export_failed",
                    "excel_rows": 0, "json_rows": 0, "missing": 0,
                    "tracker_path": tracker_path, "export_clicked": True,
                    "xlsx_path": str(xlsx_path) if xlsx_path else "",
                    "xlsx_size": xlsx_size, "headers_seen": [],
                    "json_file_exists": False,
                })
                continue

            try:
                excel_rows, diag = _read_excel_rows(xlsx_path)
                headers_seen = diag.get("headers", [])
                skipped_reasons = diag.get("skipped_reasons", {})
            except Exception as e:
                tracker_path = str(_write_missing_tracker(day_iso, []))
                out["per_day"].append({
                    "date": day_iso, "ok": False, "why": f"excel_parse_failed: {type(e).__name__}: {e}",
                    "excel_rows": 0, "json_rows": 0, "missing": 0,
                    "tracker_path": tracker_path, "export_clicked": True,
                    "xlsx_path": str(xlsx_path), "xlsx_size": xlsx_size,
                    "headers_seen": [], "json_file_exists": False,
                })
                continue

            json_rows, json_exists = _read_json_payload(day_iso)
            missing = _json_missing_vs_excel(excel_rows, json_rows)

            # ALWAYS write a tracker (even when empty) so you can audit later.
            tracker_path = str(_write_missing_tracker(day_iso, missing))

            out["processed"] += 1
            out["total_missing"] += len(missing)
            out["per_day"].append({
                "date": day_iso, "ok": True, "why": "",
                "excel_rows": len(excel_rows), "json_rows": len(json_rows),
                "missing": len(missing), "tracker_path": tracker_path,
                "export_clicked": True,
                "xlsx_path": str(xlsx_path), "xlsx_size": xlsx_size,
                "headers_seen": headers_seen, "skipped_reasons": skipped_reasons,
                "json_file_exists": json_exists,
            })

        return out

    except Exception as e:
        log.exception("[fallback] collect-missing failed")
        return {"ok": False, "error": f"{type(e).__name__}: {e}", **out}
    finally:
        try:
            if drv and not cfg.get("KEEP_BROWSER"):
                drv.quit()
        except Exception:
            pass
